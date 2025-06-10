import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Paths
CSV_PATH = './data/kinases.csv'
XLSX_PATH = './data/Kinases_Kannan_updated.xlsx'
OUTPUT_PATH = './data/Kinases_Kannan_highlighted.xlsx'

# ---------- Step 1: Read CSV uniprot IDs into a set ----------
df_csv = pd.read_csv(CSV_PATH, usecols=['uniprotid'])
df_csv = df_csv.dropna(subset=['uniprotid'])
csv_uniprots = set(df_csv['uniprotid'].astype(str).tolist())

# ---------- Step 2: Open Excel and locate "UniProt IDs" column ----------
wb = load_workbook(XLSX_PATH)
sheet_name = 'Updated Data'
if sheet_name not in wb.sheetnames:
    raise ValueError(f"Sheet '{sheet_name}' not found.")
ws = wb[sheet_name]

header = next(ws.iter_rows(min_row=1, max_row=1))
uni_col_idx = None
for cell in header:
    if cell.value == 'UniProt IDs':
        uni_col_idx = cell.column
        break
if uni_col_idx is None:
    raise ValueError("Column 'UniProt IDs' not found in header.")

# ---------- Step 3: Build a mapping from each UniProt token → Excel row numbers ----------
token_to_rows = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row_idx = row[0].row
    cell_val = row[uni_col_idx - 1].value
    if not cell_val:
        continue
    tokens = [tok.strip() for tok in str(cell_val).split(',')]
    for tok in tokens:
        if not tok:
            continue
        token_to_rows.setdefault(tok, set()).add(row_idx)

# ---------- Step 4: Determine which rows to highlight based on CSV uniprot IDs ----------
highlight_rows = set()
for up_id in csv_uniprots:
    if up_id in token_to_rows:
        highlight_rows.update(token_to_rows[up_id])

# ---------- Step 5: Apply yellow fill to matched rows ----------
yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
for row_idx in highlight_rows:
    for cell in ws[row_idx]:
        cell.fill = yellow

# ---------- Step 6: Save output ----------
wb.save(OUTPUT_PATH)
print(f'Highlighted file saved to: {OUTPUT_PATH}')