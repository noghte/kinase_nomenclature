import os
import json
from openpyxl import load_workbook

# Paths
JSON_PATH = './data/gene_to_uniprot.json'
XLSX_PATH = './data/Kinases_Kannan_updated.xlsx'
OUTPUT_PATH = './data/Kinases_Kannan_with_uniprot.xlsx'

# Sheet and column names
SHEET_NAME = 'Updated Data'
GENE_COL_NAME = 'Gene Names (human)'
NEW_COL_NAME = 'UniProt IDs'

# Load JSON mapping
with open(JSON_PATH, 'r') as f:
    entries = json.load(f)
gene_to_uniprot = {entry['value']: entry['uniprotids'] for entry in entries}

# Open workbook and select sheet
wb = load_workbook(XLSX_PATH)
if SHEET_NAME not in wb.sheetnames:
    raise ValueError(f"Sheet {SHEET_NAME} not found.")
ws = wb[SHEET_NAME]

# Identify columns
header = next(ws.iter_rows(min_row=1, max_row=1))
col_index = {cell.value: cell.column for cell in header}
if GENE_COL_NAME not in col_index:
    raise ValueError(f"Column {GENE_COL_NAME} not found in header.")

# Add new header for UniProt IDs
new_col_idx = ws.max_column + 1
ws.cell(row=1, column=new_col_idx, value=NEW_COL_NAME)

# Fill new column based on mapping
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    gene_value = row[col_index[GENE_COL_NAME] - 1].value
    if gene_value:
        key = str(gene_value).strip()
        if key in gene_to_uniprot:
            ids = gene_to_uniprot[key]
            if ids:
                ws.cell(row=row[0].row, column=new_col_idx, value=','.join(ids))

# Save updated workbook
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
wb.save(OUTPUT_PATH)
print(f"Workbook saved with UniProt IDs in new column: {OUTPUT_PATH}")